# -*- coding: utf-8 -*-
#  Copyright 2017 NTT Communications
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.

# $Rev: 1980 $
# $Ver: $
# $Date: 2019-04-05 00:22:26 +0900 (金, 05 4 2019) $
# $Author: $

""" Provides keywords for Juniper platform

*Notes:* Ignore the _self_ parameters when using those keywords.
"""

import os,re
import codecs
import json
import jinja2
import time
import netsnmp
import shutil
import Common
from datetime import datetime
from datetime import timedelta
import robot.libraries.DateTime as DateTime
from robot.libraries.BuiltIn import BuiltIn
from robot.libraries.OperatingSystem import OperatingSystem
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment


def get_version(self):
    """ Returns router version information
    """ 

    result = self._vchannel.cmd('show version | no-more')
    return result

def get_current_datetime(self,time_format='%H:%M:%S',delta_time='0s',dir='+',**kwargs):    
    """ Returns the current date time with vendor format
    ``delta_time`` will be added or subtracted to current time, default is ``0s``

    ``time_format`` decides the time part of the output.
    Example result are :
    | May 24 04:14:25 
    | May  4 04:14:25
    *Note:* The date part is padded by space, and the result is allways 15 characters
    """
 
    delta = DateTime.convert_time(delta_time)
    if dir == '-':
        current_time = datetime.now() - timedelta(seconds=delta)
    else:
        current_time = datetime.now() + timedelta(seconds=delta)
    padded_day = datetime.strftime(current_time,"%d").rjust(2)
    format = "%%b %s %s" % (padded_day,time_format)
    result  = datetime.strftime(current_time,format)
    
    BuiltIn().log("Got current datetime and convert to Juniper format")
    return result


def number_of_ospf_neighbor(self,state="Full",cmd='show ospf neighbor'):
    """ Returns number of OPSF neighbors with status ``state``
    """
    output  = self._vchannel.cmd(cmd).lower()
    count   = output.count(state.lower())

    BuiltIn().log("Number of OSPF neighbors in `%s` state is %d" % (state,count))
    return count


def number_of_ospf3_neighbor(self,state="Full",cmd='show ospf3 neighbor'):
    """ Returns number of OPSFv3 neighbors with status ``state``
    """
    output  = self._vchannel.cmd(cmd)
    count   = output.count(state)

    BuiltIn().log("Number of OSPF neighbors in `%s` state is %d" % (state,count))
    return count


def number_of_bgp_neighbor(self,state="Established",cmd='show bgp neighbor | match "Type"'):
    """ Returns number of BGP neighbor in ``state`` state
    """
    output  = self._vchannel.cmd(cmd).lower()
    count   = output.count(state.lower())

    BuiltIn().log("Number of BGP neighbors in `%s` state is %d" % (state,count))
    return count


def enable_interface(self,intf):
    """ Enables an interface ``intf``
    """
 
    self._vchannel.cmd("configure")
    self._vchannel.cmd("delete interface " + intf + " disable")
    self._vchannel.cmd("commit")
    self._vchannel.cmd("exit")

    BuiltIn().log("Enabled interface `%s`" % (intf))


def disable_interface(self,intf):
    """ Disables an interface ``intf``
    """
    
    self._vchannel.cmd("configure")
    self._vchannel.cmd("set interface " + intf + " disable")
    self._vchannel.cmd("commit")
    self._vchannel.cmd("exit")

    BuiltIn().log("Disabled interface `%s`" % (intf))


def flap_interface(self,intf,time_str='10s'):
    """ Simulates an interface flap for interface ``intf``

    Disables the interface and wait for a while before turning it up again
    """

    self._vchannel.cmd("configure")
    self._vchannel.cmd("set interface " + intf + " disable")
    self._vchannel.cmd("commit")

    time.sleep(DateTime.convert_time(time_str))

    self._vchannel.cmd("delete interface " + intf + " disable")
    self._vchannel.cmd("commit")
    self._vchannel.cmd("exit")

    BuiltIn().log("Flapped interface `%s`" % (intf))
    

def get_cli_mode(self):
    """ Returns current mode of the CLI.

    Return value is ``config`` for configuration mode or ``command`` for command
    mode
    """
    result = ""
    output = self._vchannel.cmd('')
    if "#" in output: result = "config"
    if ">" in output: result = "command"

    BuiltIn().log("CLI mode is `%s`" % (result))
    return result


def push_config(self,mode='set',config_file='', \
                        pre_config=None, \
                        pos_config=None, \
                        confirm='0s',vars='',err_match='(error:|unknown command:)'):
    """ Pushes configuration directly to the router

    Usable ``mode`` is ``set``, ``override``, ``merge`` and ``replace``

    ``set`` mode uses configuration that contains ``set`` command.
    Mode ``override``, ``merge`` and ``replace`` use ordinary JunOS configuration file with appropriate mode.
    ``config_file`` is a configuration file inside the ``config`` folder of the
    current test case.
    
    Config file could includes jinja2 template. The template will be evalued
    with `LOCAL`, `GLOBAL` and varibles defined by `vars`. The `vars` has the
    format: var1=value1,var2=value2 ...
    
    If the loading has no error that match the ``error_match``, the
    configuration will be commited.

    The keywordl waits for ``confirm`` seconds before rollback the commited configuration. A
    zero value indicates an immediatly commit

    `pre_config` and `pos_config` is extra configuration commands separated by
    `;` that would be excuted before and after the configuration is loaded. It could
    be used to add extra firewall rules to the router.

    *Note*: by default the keyword will activate the SSH service on the router
    if the service is not activated and disable the service after loading the
    configuration.
    """
    # validate mode
    if not mode in ['override','merge','replace','set']:
        raise Exception("Invalid ``mode``. ``mode`` should be ``set``,``override``,``merge``,``replace``")

    # copy file, assuming current mode is command mode
    cli_mode = self.get_cli_mode()
    if cli_mode == "config": self._vchannel.cmd('exit')

    confirm_time = int(DateTime.convert_time(confirm) / 60) # minute
    current_services = self.cmd('show configuration system services')
    self.cmd('configure',prompt='# ')
    if 'ssh' not in current_services:
        self.cmd('set system service ssh',prompt='# ')
        if pre_config is not None:
            for item in pre_config.split(';'): self.cmd(item,prompt='# ')
    self.cmd('commit synchronize')

    # prepare the configuration file
    folder              = os.getcwd() + '/config/'
    file_path           = os.getcwd() + '/config/' + config_file
    file_path_replace   = os.getcwd() + '/tmp/' + config_file + '.tmp'
    # jinja2 process
    loader=jinja2.Environment(loader=jinja2.FileSystemLoader(folder)).get_template(config_file)
    render_var = {'LOCAL':Common.LOCAL,'GLOBAL':Common.GLOBAL}
    for pair in vars.split(','):
        info = pair.split("=")
        if len(info) == 2:
            render_var.update({info[0].strip():info[1].strip()})
    compiled_config = loader.render(render_var)
    with codecs.open(file_path_replace,'w','utf-8') as f: 
        f.write(compiled_config)
    BuiltIn().log('Compiled and wrote configuration to `%s`' % file_path_replace)
    file_path_replace = file_path_replace.replace('(','\(').replace(')','\)')

    # load the configuration
    _user = self._vchannel.get_current_channel()['auth']['user'] 
    _pass = self._vchannel.get_current_channel()['auth']['pass'] 
    _ip   = self._vchannel.get_current_channel()['ip'] 
    cmd = "sshpass -p %s scp -o StrictHostKeyChecking=no -o UserKnownHostsFile=/dev/null %s %s@%s:/var/tmp/%s" % (_pass,file_path_replace,_user,_ip,config_file)
    rc,output = OperatingSystem().run_and_return_rc_and_output(cmd)

    # cleanup additional configuration
    if 'ssh' not in current_services:
        self.cmd('delete system services ssh')
    if pos_config is not None:
        for item in pos_config.split(';'): self.cmd(item,prompt='# ')
    self.cmd('commit synchronize')

    # load the configuration
    if rc != 0:
        self.cmd('exit')
        msg = "ERROR: error while pushing configuation to router;\n%s" % output
        BuiltIn().log(msg)
        raise(Exception(msg))
    else:
        output = self.cmd("load " + mode + " /var/tmp/" + config_file)

        # check ouput
        if re.search(err_match, output, re.MULTILINE):    
            self.cmd("rollback 0")
            self.cmd("exit")
            msg = "ERROR: An error happened while loading the config. Output: `%s`" % output
            BuiltIn().log(msg)
            BuiltIn().log('output:')
            BuiltIn().log(output)
            raise Exception(msg)

    if confirm_time == 0:
        output = self.cmd('commit synchronize')
    else: 
        output = self.cmd("commit confirmed %s" % (confirm_time))

    # check output 
    if re.search(err_match, output, re.MULTILINE):
        self.cmd("rollback 0")
        self.cmd("exit")
        msg = "ERROR: An error happened while committing the change so I rolled it back"
        BuiltIn().log(msg)
        BuiltIn().log('output:')
        BuiltIn().log(output)
        raise Exception(msg)
    else:
        self.cmd('exit')

    BuiltIn().log("commit result is: " + output)
    BuiltIn().log("Loaded config with ``%s`` mode and confirm time %s" % (mode,confirm_time))
    


def load_config(self,mode='set',config_file='',confirm='0s',vars='',err_match='(error:|unknown command:)'):
    """ Loads configuration to a router. 
    Usable ``mode`` is ``set``, ``override``, ``merge`` and ``replace``

    ``set`` mode uses configuration that contains ``set`` command.
    Mode ``override``, ``merge`` and ``replace`` use ordinary JunOS configuration file with appropriate mode.
    ``config_file`` is a configuration file inside the ``config`` folder of the
    current test case.
    
    Config file could includes jinja2 template. The template will be evalued
    with `LOCAL`, `GLOBAL` and varibles defined by `vars`. The `vars` has the
    format: var1=value1,var2=value2 ...
    
    If the loading has no error that match the ``error_match``, the
    configuration will be commited.

    The keywordl waits for ``confirm`` seconds before rollback the commited configuration. A
    zero value indicates an immediatly commit
    """
    # validate mode
    if not mode in ['override','merge','replace','set']:
        raise Exception("Invalid ``mode``. ``mode`` should be ``set``,``override``,``merge``,``replace``")

    # copy file, assuming current mode is command mode
    cli_mode = self.get_cli_mode()
    if cli_mode == "config": self._vchannel.cmd('exit')

    server      = Common.GLOBAL['default']['robot-server']
    password    = Common.GLOBAL['default']['robot-password']
    # the original config is in ./config folder
 
    folder              = os.getcwd() + '/config/'
    file_path           = os.getcwd() + '/config/' + config_file
    file_path_replace   = os.getcwd() + '/tmp/' + config_file + '_tmp'

    # jinja2 process
    loader=jinja2.Environment(loader=jinja2.FileSystemLoader(folder)).get_template(config_file)
    render_var = {'LOCAL':Common.LOCAL,'GLOBAL':Common.GLOBAL}
    for pair in vars.split(','):
        info = pair.split("=")
        if len(info) == 2:
            render_var.update({info[0].strip():info[1].strip()})
    compiled_config = loader.render(render_var)

    with codecs.open(file_path_replace,'w','utf-8') as f: 
        f.write(compiled_config)
    file_path_replace = file_path_replace.replace('(','\(').replace(')','\)')
    cmd = 'file copy robot@%s:\'//%s\' /var/tmp/%s' % (server,file_path_replace,config_file)

    output = self._vchannel.cmd(cmd,prompt="\(yes/no\)\? |password: ")
    if "yes/no" in output:
        output = self.cmd("yes",prompt='password: ')
    if "password:" in output:
        output = self.cmd(password)

    confirm_time = int(DateTime.convert_time(confirm) / 60) # minute

    if not '100%' in output:
        raise Exception("ERROR: error while copying config file `%s`" % config_file)
    else:
        self.cmd("configure")
        output = self.cmd("load %s /var/tmp/%s" % (mode,config_file))
        if re.search(err_match, output, re.MULTILINE):    
            self.cmd("rollback 0")
            self.cmd("exit")
            msg = "ERROR: An error happened while loading the config:\n%s" % output
            BuiltIn().log(msg)
            raise(Exception(msg))

    if confirm_time == 0:
        output = self.cmd('commit synchronize')
    else: 
        output = self.cmd("commit confirmed %s" % (confirm_time))

    # check output 
    if re.search(err_match, output, re.MULTILINE):
        self.cmd("rollback 0")
        self.cmd("exit")
        msg = "ERROR: An error happened while committing the change so I rolled it back\n%s" % output
        BuiltIn().log(msg)
        raise(Exception(msg))
    else:
        self.cmd('exit')
    BuiltIn().log("commit result is: " + output)
    BuiltIn().log("Loaded config with ``%s`` mode and confirm time %s" % (mode,confirm_time))


def copy_file(self,src_path,filename=None,pre_config=None,pos_config=None):
    """ Copies a file from router

    Parameters:
    - `src_path`: a absolute path insides the router
    - `filename`: a file name under ``result`` folder
    """
    # get current service and add ssh if it is necessary
    current_services = self.cmd('show configuration system services')
    if 'ssh' not in current_services:
        self.cmd('configure',prompt='# ')
        self.cmd('set system service ssh',prompt='# ')
        if pre_config is not None:
            for item in pre_config.split(';'): self.cmd(item,prompt='# ')
        self.cmd('commit synchronize and-quit')

    # prepare filepath
    if filename is None:
        _filename = os.path.basename(src_path)
    else:    
        _filename = filename
    dst_path  = '%s/%s' % (Common.get_result_path(),_filename)
    dst_path = dst_path.replace('(','\(').replace(')','\)')

    #
    _user = self._vchannel.get_current_channel()['auth']['user'] 
    _pass = self._vchannel.get_current_channel()['auth']['pass'] 
    _ip   = self._vchannel.get_current_channel()['ip'] 
    cmd = 'sshpass -p %s scp -o StrictHostKeyChecking=no -o UserKnownHostsFile=/dev/null %s@%s:%s %s' % (_pass,_user,_ip,src_path,dst_path)
    rc,output = OperatingSystem().run_and_return_rc_and_output(cmd)

    BuiltIn().log(output)
   
    # cleanup additional configuration
    if 'ssh' not in current_services:
        self.cmd('configure',prompt='# ')
        self.cmd('delete system services ssh')
        if pos_config is not None:
            for item in pos_config.split(';'): self.cmd(item,prompt='# ') 
        self.cmd('commit synchronize and-quit')
  
    if  rc != 0:
        BuiltIn().log(output)
        raise(Exception("ERROR: error while copy a file with error:\n%s" % output))
  
    BuiltIn().log('Copied file from `%s` to `%s`' % (src_path,filename))



def get_file(self,src_file,dst_file=None):
    """ Gets a file from router

    Parameters:
    - `src_file`: a absolute path insides the router
    - `dst_file`: a file name under ``result`` folder

    if `dst_file` is not defined, it will be the filename of the `src_file`.

    The keyword copy the specific file *FROM* the router to the RENAT server
    """
 
    cli_mode = self.get_cli_mode()
    if cli_mode == "config": self._vchannel.cmd('exit')
    self._vchannel.stop_screen_mode()
   
    server      = Common.GLOBAL['default']['robot-server']
    password    = Common.GLOBAL['default']['robot-password']


    if dst_file is None:
        filename = os.path.basename(src_file)
        tmp_path    = os.getcwd() + '/tmp/%s' % filename
        dest_path   = os.getcwd() + '/' + Common.get_result_folder() + filename
    else:    
        tmp_path    = os.getcwd() + '/tmp/' + dst_file
        dest_path   = os.getcwd() + '/' + Common.get_result_folder() + '/' + dst_file

    dest_path = dest_path.replace('(','\(').replace(')','\)')
    tmp_path = tmp_path.replace('(','\(').replace(')','\)')

    self._vchannel.cmd('start shell',prompt='% ')
    self._vchannel.cmd('chmod g+rw %s' % src_file, prompt='% ')
    self._vchannel.cmd('exit')

    cmd = 'file copy %s robot@%s:\'//%s\'' % (src_file,server,tmp_path)
    output = self._vchannel.cmd(cmd,prompt="\(yes/no\)\? |password: ")

    if "yes/no" in output:
        output = self._vchannel.cmd("yes",prompt='password: ')
    if "password:" in output:
        output = self._vchannel.cmd(password)
    if "error" in output:
        BuiltIn().log("ERROR:")
        BuiltIn().log(output)
        raise Exception("ERROR:" + output) 

    # copy config from temp folder to result folder
    shutil.copy(tmp_path,dest_path)
    Common.change_mod(dest_path,'0775',False) 
    
    BuiltIn().log("Get the file `%s` from node `%s`" % (src_file,self._vchannel.current_name))


def copy_config(self,dst_name=None):
    """ Gets the configuration file of the router

    This keyword directly copy the configuration file from the router
    """
    return self.copy_file('/config/juniper.conf.gz',dst_name)

def get_config(self,dst_name=None):
    """ Gets the current configuration file of the router to current ``result``
    folder.

    Default ``dst_name`` is ``juniper.conf.gz``
    This keyword push the configuration from *FROM* the router to the RENAT
    server
    """
    return self.get_file('/config/juniper.conf.gz',dst_name)


def link_status(self,if_name):
    """ Returns link physical status as string (aka: "up down", "up up")
    """
   
    result = "" 
    output = self._vchannel.cmd("show interface %s terse | grep \"%s \"" % (if_name,if_name))
    tmp = re.split('\s+',output.split("\n")[0].strip())
    if len(tmp) == 3:
        result = tmp[1] + " " + tmp[2]
        BuiltIn().log("Got link status of `%s`: %s %s" % (if_name,tmp[1],tmp[2]))
    else:
        raise Exception("Error while getting link status of `%s`" % if_name)

    return result


def get_route_number(self,table='inet.0'):
    """ Returns number of active route in the ``table``

    ``table`` could be ``inet.0`` or ``inet.6``
    """
    
    result = ""
    output = self._vchannel.cmd("show route summary table %s" % table)
    for line in output.split("\n"):
        match = re.match(r"%s: .* \((.+) active," % table, line)
        if match:
            result = int(match.group(1))
    BuiltIn().log("Got %d routes from `%s`" % (result,table)) 
    
    return result 


def get_intf_addr(self,intf_name,family='inet'):
    """ Returns the tuple of address and netmask of an interface

    ``family`` should be ``inet`` or ``inet6``
    If the address is not set, ``('','')`` will be returned.
    """
    
    output  = self._vchannel.cmd("show interface %s terse | match %s" % (intf_name,family))
    line = output.split('\n')[0]    # first line
    try:
        address = tuple(line.split()[-1].split('/'))
    except IndexError:
        address = ('','')
    BuiltIn().log("Interface address is `%s`, netmask is `%s`" % (address[0],address[1]))
    return address

def get_chassis_serial(self):
    """ Returns the serial number of the chassis
    """

    output = self._vchannel.cmd("show chassis hardware | match Chassis")
    line = output.split('\n')[0]    # first line
    tmp = line.split()
    if len(tmp) > 2:
        result = tmp[1]
    else:
        result = ''
    BuiltIn().log("Got the serial number: %s" % (result))
    return result


def create_best_path_select_data(self,route_content,output_excel='best.xlsx'):
    """ Creates the matrix of best path selection 
    
    Provides the test described in  `smb://10.128.3.91/SharePoint01/31_VerificationRoom/31_13_検証環境セット/BGP-Best-Path-SelectionのAll-in-One設定_20161118改良/`
    The test uses predefined Ixia config and follows predefined steps
    """

    wb = openpyxl.Workbook()
    wb.guess_types = True
    tmp = output_excel.split('.')
    sheet_name = tmp[0]
    ws = wb.create_sheet(sheet_name,0)

    # adding column name
    ws.append( ['INTF','Prefix','ProtoPref','LocalPref','AS_PATH','Origin','MED','Protocol','I/E','IGP','Age','RouterID','CLLength','Peer','Win/Loose','Reason'])
   
    # sparsing route information 
    route_info_list=re.findall(r"(\S+/.{1,3}) (?:.*?entr.*?announced.*?)\r\n(( {8}[ *](\S+) *Pref.*?\r\n( {16}.*?\r\n)+)+)\r\n",route_content,re.DOTALL)

    line = 1 # excel start at number 1
    for route_info in route_info_list:
        route = route_info[0]
        proto_list = re.findall(r" {8}([ *])(\S+) *Preference: (\S+?)(?:/\S+){0,1}\r\n(( {16}.*?\r\n)+)",route_info[1],re.DOTALL)
        for proto_info in proto_list:
            if proto_info[0] == '*':
                win = 'win'
                if line > 1:
                    ws.append( ['','','','','','','','','','','','','','','',''])
                    line = line + 1
            else:
                win = 'loose'
            proto   = proto_info[1]
            pp      = proto_info[2]
    
            match_intf = re.search(r" via (\S+),", proto_info[3],re.MULTILINE)
            if match_intf:
                intf_index  = int(match_intf.group(1).split('.')[1])
                via_intf = chr(ord('A')+intf_index-1)
            else:
                via_intf = '-'
    
            match_localpref = re.search(r"Localpref: (\S+)", proto_info[3],re.MULTILINE)
            if match_localpref: local_pref = match_localpref.group(1)
            else:               local_pref = '-'
    
            match_as_path = re.search(r"AS path: (.*?)\r\n", proto_info[3],re.MULTILINE)
            if match_as_path:
                as_list = match_as_path.group(1).split()
                as_path = len(as_list) - 1 # the last is the origin
                if as_list[-1] == 'E':
                    origin = 'EGP'
                else:
                    origin = 'IGP'
            else:
                as_path = '-'
                origin = '-'
    
            match_med = re.search(r"Metric: (\S+) ", proto_info[3],re.MULTILINE)
            if match_med:
                med = match_med.group(1)
            else:
                med = '-'
    
            match_type = re.search(r"Local AS:  (\S+) Peer AS: (\S+)",proto_info[3],re.MULTILINE)
            if match_type:
                if match_type.group(1) == match_type.group(2):
                    type = 'IBGP'
                else:
                    type = 'EBGP'
            else:
                type = '-'
    
    
            match_igp = re.search(r"Metric2: (\S+)",proto_info[3],re.MULTILINE)
            if match_igp:
                igp = match_igp.group(1)
            else:
                igp = '-'
    
            match_router_id = re.search(r"Router ID: (\S+)",proto_info[3],re.MULTILINE)
            if match_router_id:
                router_id = match_router_id.group(1)
            else:
                router_id = '-'
    
            match_cll = re.search(r"Cluster list:  (.*)\r\n",proto_info[3],re.MULTILINE)
            if match_cll:
                cll = len(match_cll.group(1).split())
            else:
                cll = '0'
    
            match_peer = re.search(r"Source: (\S+)",proto_info[3],re.MULTILINE)
            if match_peer:
                peer = match_peer.group(1)
            else:
                peer = '-'
    
#            match_age = re.search(r"Age: (\S+)",proto_info[3],re.MULTILINE)
#            if match_age:
#                age = reduce(lambda x, y: x*60+y, [int(i) for i in match_age.group(1).split(':')])
#            else:
#                age = '-'
            match_age = re.search(r"Age: (.*?) (?:\r\n|    )", proto_info[3],re.MULTILINE)
            if match_age:
                m = re.match(r"(\d+w){0,1}(\d+d){0,1} {0,1}(\S+)",match_age.group(1))
                age = 0
                if m.group(1): age = age + int(m.group(1)[:-1])*7*24*60*60
                if m.group(2): age = age + int(m.group(2)[:-1])*24*60*60
                age = age + sum(int(x) * 60 ** i for i,x in enumerate(reversed(m.group(3).split(":"))))
            else:
                age = '-'
    
            match_reason = re.search(r"Inactive reason: (.+)\r\n",proto_info[3],re.MULTILINE)
            if match_reason:
                reason = match_reason.group(1)
            else:
                reason = '-'
    
    
            BuiltIn().log("    Added row: \
            (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)" %
            (via_intf,route,pp,local_pref,as_path,origin,med,proto,type,igp,age,router_id,cll,peer,win,reason))
    
            row = [via_intf,route,pp,local_pref,as_path,origin,med,proto,type,igp,age,router_id,cll,peer,win,reason]
            ws.append(row)
            line = line + 1
   
    # coloring style
    # HeaderFill  = PatternFill(start_color='ffff00',end_color='ffff00',fill_type='solid')
    RedFont     = Font(color=colors.RED)
    BlueFont    = Font(color=colors.BLUE)
    #for cell in ws[1]:
    #    cell.fill = HeaderFill
    #for cell in ws['O']:
    #    if cell.value=='win': cell.font = RedFont
    
    for cell in reversed(ws['P']):
        row = int(cell.row)
        if cell.value == 'Update source':
            ws['N%d' % row].font = BlueFont
            ws['N%d' % (row - 1)].font = RedFont
        if cell.value == 'Cluster list length':
            ws['M%d' % row].font = BlueFont
            ws['M%d' % (row - 1)].font = RedFont
        if cell.value == 'Router ID':
            ws['L%d' % row].font = BlueFont
            ws['L%d' % (row - 1)].font = RedFont
        if cell.value == 'Active preferred':
            ws['K%d' % row].font = BlueFont
            ws['K%d' % (row - 1)].font = RedFont
        if cell.value == 'IGP metric':
            ws['J%d' % row].font = BlueFont
            ws['J%d' % (row - 1)].font = RedFont
        if cell.value == 'Interior > Exterior > Exterior via Interior':
            ws['I%d' % row].font = BlueFont
            ws['I%d' % (row - 1)].font = RedFont
        if cell.value == 'Route Metric or MED comparison':
            ws['H%d' % row].font = BlueFont
            ws['H%d' % (row - 1)].font = RedFont
        if cell.value == 'Always Compare MED':
            ws['G%d' % row].font = BlueFont
            ws['G%d' % (row - 1)].font = RedFont
        if cell.value == 'Origin':
            ws['F%d' % row].font = BlueFont
            ws['F%d' % (row - 1)].font = RedFont
        if cell.value == 'AS path':
            ws['E%d' % row].font = BlueFont
            ws['E%d' % (row - 1)].font = RedFont
        if cell.value == 'Local Preference':
            ws['D%d' % row].font = BlueFont
            ws['D%d' % (row - 1)].font = RedFont
        if cell.value == 'Route Preference':
            ws['C%d' % row].font = BlueFont
            ws['C%d' % (row - 1)].font = RedFont
    
    # adjust column width
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['O'].width = 12
    ws.column_dimensions['P'].width = 32
    
    tab = Table(displayName="Table1",ref="A1:P%d" % line)
    style = TableStyleInfo(name="TableStyleMedium9",showFirstColumn=False,showLastColumn=False,showRowStripes=False,showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # top = Border(top=border.top)
    rows = ws["A1:P%d" % line]
    for row in rows:
        for cell in row:
            cell.alignment = Alignment(horizontal="right")
    
    ws.append([])
    ws.append(['※1 Protocol Preference の値はルータの import policy で変更'])
    ws.append(['※2 ここでは Always Compare MED 有りの試験のみ実施。'])
    ws.append(['※3 1.6.0.0/staticの場合、ルータにて PP 170 の static経路を設定。ただし、正常に試験できているか不明。'])
    ws.append(['※4 全経路配信後、IF-B から配信しているこの経路(1.9.0.0) だけ一回フラップさせるた。'])

    save_path = os.getcwd() + '/' + Common.get_result_folder() + '/' + output_excel
    wb.save(save_path) 
    BuiltIn().log("Created the best path select matrix")

def _check_interface_output(output, link):
    '''Used to check the output from commands which target a specific interface name, 
    validating the output before handing it off for further processing.

    Checks for empty return, or common errors such as interface not found
    '''
    if 'not found' in output:
        return ('%s not found' % link)
    else:
      return True

def _check_if_output_empty(output):
    '''Used to check the output from commands to make sure the return is not blank/Empty, 
    validating the output before handing it off for further processing.

    Checks for empty return represented by the prompt and a new line.

    Returns False if the test fails (blank output)
    True if the test passes (has output)
    '''
    config_lines = output.rstrip().split('\n')
    conlen = len(config_lines)
    if 'error' in output:
        return False
    if 'master' in output or 'backup' in output:
        if conlen > 3:
            return True
        else:
            if re.match(r'^.*>$',config_lines[2]):
                return False
            else:
                return True
    else:
        if conlen > 2:
            return True
        else:
            if re.match(r'^.*>$',config_lines[1]):
                return False
            else:
                return True
        
def _check_link_status_description(output,description):
    if description in output:
        return True
    else:
        raise Exception('desciption not in output')

def _check_link_status_state(output,status):
    if status == 'up':
        if 'Physical link is Up' in output:
            return True
        else:
            raise Exception('Link not in Desired State - Should be %s' %status)
    elif status == 'down':
        if 'Physical link is Down' in output:
            return True
        else:
            raise Exception('Link not in Desired State - Should be %s' %status)
    else:
        raise Exception('Valid link state not provided')

def _verify_route_engine_status(output):
    """ Attempts to verify that RE0 has the mastership.
    Failure indicates either RO0 is missing or another RE has the master role.
    """ 
    masterstatus = re.search(r'Current state.*(Master)',output)
    if masterstatus != None:
        return True
    else:
        return False
    

def get_interface_information(self,state='all'):
    '''Captures the information about interfaces in a defined state.

    If no state specified it will capture all interfaces

    Get Interface Information | state=up
    '''
    device = self._vchannel.current_name

    if state == 'up':
        output = self._vchannel.cmd('show interfaces terse | match up.*up | no-more')
        if _check_if_output_empty(output) != True:
	        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    elif state == 'down':
        output = self._vchannel.cmd('show interfaces terse | match down.*down | no-more')
        if _check_if_output_empty(output) != True:
	        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    elif state == 'all':
        output = self._vchannel.cmd('show interfaces terse | no-more')
        if _check_if_output_empty(output) != True:
	        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    else:
        raise Exception('Invalid Arguments Provided. Please use "up" or "down"')

def check_chassis_environment(self):
    '''Checks the chassis environment at a high level, validating if any errors exist within the environment

    Check Chassis Environment
    '''
    device = self._vchannel.current_name

    output = self._vchannel.cmd('show chassis environment')
    
    if _check_if_output_empty(output) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    if 'Check' in output:
        BuiltIn().log('\n   Error In Environment',console=True)
        raise Exception('Error In Environment')
    else:
        BuiltIn().log('\n   Chassis Environment Acceptable',console=True)

#Check the output differences noted in the test and what i saw on CR
def check_chassis_environment_fans(self):
    '''Validates the status of the fans.
    '''
    output = self._vchannel.cmd('show chassis environment | grep fan')
    if 'Check' in output:
        raise Exception('Error In Environment')
    else:
        BuiltIn().log('\n   Chassis Environment Acceptable',console=True)

def check_chassis_pem_modules(self):
    '''Checks the status of all Power Entry Modules (PEM) for any errors exist within the environment.

    Check Chassis PEM Modules
    '''
    device = self._vchannel.current_name
    output = self._vchannel.cmd('show chassis environment pem | no-more')

    if _check_if_output_empty(output) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    if 'Check' in output:
        raise Exception('Error In Environment')
    else:
        BuiltIn().log('\n   Chassis Environment Acceptable',console=True)

def check_chassis_craft_interface_alarm_status(self):
    '''Checks the alarm status of the craft interface.

    Check Chassis Craft Interface Alarm Status
    '''
    device = self._vchannel.current_name

    redalarm = self._vchannel.cmd('show chassis craft-interface | match "Red LED"')
    if _check_if_output_empty(redalarm) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    yellowalarm = self._vchannel.cmd('show chassis craft-interface | match "Yellow LED"')
    if _check_if_output_empty(yellowalarm) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    majorrelay= self._vchannel.cmd('show chassis craft-interface | match "Major relay"')
    if _check_if_output_empty(majorrelay) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    minorrelay = self._vchannel.cmd('show chassis craft-interface | match "Minor relay"')
    if _check_if_output_empty(minorrelay) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    if '*' in (redalarm , yellowalarm, majorrelay, minorrelay): #Specifically check for failed LEDs
        raise Exception

def verify_chassis_alarm_status(self):
    ''' Verifies the alarm status of the chassis.

    Fails if any alarms are currently active.

    Check Chassis Alarm Status
    '''
    device = self._vchannel.current_name

    output = self._vchannel.cmd('show chassis alarms')

    if _check_if_output_empty(output) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    if 'No alarms currently active' not in output:
        raise Exception('Alarms currently active on device')

def check_chassis_routing_engine(self):
    ''' Checks the routing engine has no failure the alarm status of the chassis.

    Check Chassis Routing Engine
    '''
    device = self._vchannel.current_name

    output = self._vchannel.cmd('show chassis craft-interface')

    if _check_if_output_empty(output) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    if 'Fail              .    .' in output: #Specifically check for failed LEDs
        pass
    else:
        raise Exception('Error In Environment')

def check_chassis_power(self):
    ''' Checks the power status of the chassis.

    Check Chassis Power
    '''
    device = self._vchannel.current_name

    output = self._vchannel.cmd('show chassis power')
    
    if _check_if_output_empty(output) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    if 'Check' in output:
        raise Exception('Error In Environment')
    else:
        BuiltIn().log('\n   Chassis Environment Acceptable',console=True)

def clear_bgp_routes(self,table='inet.0'):
    ''' Clears the specified route table of bgp routes

    Clear BGP Routes | table=inet.0
    '''
    self._vchannel.cmd('clear bgp table %s' % table)    

def check_lfm_status(self,link='default',status='Down',state='Down'):
    ''' Checks the Link Fault Management Status of the specified interfaces.
    If no arguments are specified it will simply log the output to the console.

    If arguments are specified in the form of Link, Status and State then it will validate these parameters

    Check LFM Status | link= ae1 | status up | state=up
    '''
    def _check_int_lfm_status(output,link=None,expected_status=None,expected_state=None):
        #Create errors for appending into
        errors = []

        #Get Details from the output string
        current_status = re.findall(r'Status: (\w+)', output)
        current_state = re.findall(r'state: (\w+.*)', output)
        #Validate the expected status vs the actual current
        if str(current_status[0]).strip('\r') in expected_status:
            pass
        else:
            errors.append('%s Link Fault Management not configured as expected.  Got - %s , Expected - %s' %(link,current_status,expected_status))
        if str(current_state[0]).strip('\r') in expected_state:
            pass
        else:
            errors.append('%s Link Fault Management not configured as expected.  Got - %s , Expected - %s' %(link,current_state,expected_state))
        #If there have been raised errors return them to the parent function for raising (Prevents needing to pass hostname, Centeralizes exceptions for function)
        if len(errors):
            return errors
        else:
            return True

    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name

    #Gather info based on Arguments
    if link!='default':
        output = self._vchannel.cmd('show oam ethernet link-fault-management %s' % (link))
        if _check_if_output_empty(output) != True:
            errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)
            #Check if any errors where raised as part of the process and raise them.
            if len(errors) > 0:
                raise Exception(errors) 
    elif link=='default':
        output = self._vchannel.cmd('show oam ethernet link-fault-management brief')
        if _check_if_output_empty(output) != True:
            errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)
            #Check if any errors where raised as part of the process and raise them.
            if len(errors) > 0:
                raise Exception(errors) 

    #Perform actions on gathered info based on arguments
    if link=='default':
        BuiltIn().log(output,console=True)
    elif link!='default':
        validation = _check_int_lfm_status(output,link,status,state)
        if validation != True:
            BuiltIn().log(validation, console=True)
            raise Exception(validation)

def check_link_error_status(self,link='default'):
    ''' Checks specified interface link for any errors, Having Any errors will produce a failure

    Check Link Error Status | link=ae1.0
    '''
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name

    if link == 'default':
        errors.append('%s No Link Specified, Please Check Test Script' % link)
    if len(errors) > 0:
        raise Exception(errors)
    
    output = self._vchannel.cmd('show interfaces %s extensive | display json' % link)

    if _check_if_output_empty(output) != True:
        errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)
    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)
    
    trimmed_output = re.findall(r'^(.*)(?<=).{master}|{backup}|.*>$', output,re.M | re.DOTALL)
    #print(trimmed_output[0])
    if len(trimmed_output) < 1:
        raise Exception('No Valid Output Returned For %s' % link)
    if '[' not in trimmed_output[0]:
        raise Exception('No Valid Output Returned For %s' % link)
    loadedjson = json.loads(trimmed_output)
    #loadedjson = json.loads(output)
    input_error_expression = parse('interface-information[*].physical-interface[*].input-error-list[*].*[*].data')
    input_error = input_error_expression.find(loadedjson)
    output_error_expression = parse('interface-information[*].physical-interface[*].output-error-list[*].*[*].data')
    output_error = output_error_expression.find(loadedjson)
    for error in input_error:
        if int(error.value) > 0:
            errors.append('%s has an error value of %s' % (error.context.context.path, error.value))
        else:
            pass
    for error in output_error:
        if str(error.context.context.path) in ('carrier-transitions'):
        #if 'carrier-transitions' in str(error.context.context.path):
            pass
        else:
            if int(error.value) > 0:
                errors.append('%s has an error value of %s' % (error.context.context.path, error.value))
            else:
                pass
    if len(errors) > 0:
        raise Exception(errors)

def check_mtu(self,link,link_type,expected_mtu):
    '''Used to check the MTU on the specified link
    No arguments are optional all must be provided when issuing the command

    Example - Check MTU | xe-0/0/0 | physical | 1514
            - Check MTU | xe-0/0/0.0 | logical | 1514

    incorrect use of physical and logical can lead to unexpected results/failures.

    Check will fail if MTU returned from interface does not match the MTU specified in the argument
    '''
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name

    output = self._vchannel.cmd('show interfaces %s' % link)
    if _check_if_output_empty(output) != True:
        errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)
		
    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)

    validate_interface_command_output = _check_interface_output(output, link)
    if validate_interface_command_output != True:
        # raise Exception ('Error in command output - %s' % validate_interface_command_output)
        errors.append('Error in command output - %s' % validate_interface_command_output)
    elif validate_interface_command_output == True:
        if 'physical' in link_type:
            if '.' in link:
                print('Link not Physical')
                errors.append('Link %s is not a valid phsyical interface' % (link))
            link_mtu = re.findall(r'MTU: (\d+),',output)
            print(link_mtu[0])
            if int(link_mtu[0]) != int(expected_mtu):
                print('Physical Danger Zone')
                errors.append('Link %s MTU is %d not %d as expected' % (link, int(link_mtu[0]), int(expected_mtu)))
            elif int(link_mtu[0]) == int(expected_mtu):
                print('Physical Not Danger Zone')
                return True
            else:
                errors.append('Link %s Unexpected error' % link)
        elif 'logical' in link_type:
            if '.' not in link:
                print('Link not logical')
                errors.append('Link %s is not a valid subinterface' % (link))
            else:
                link_mtu = re.findall(r'MTU: (\d+),',output)
                print(link_mtu[0])
                if int(link_mtu[0]) != int(expected_mtu):
                    print('Logical Danger Zone')
                    errors.append('Link %s MTU is %d not %d as expected' % (link, int(link_mtu[0]), int(expected_mtu)))
                elif int(link_mtu[0]) == int(expected_mtu):
                    print('Logical Not Danger Zone')
                    return True
                else:
                    errors.append('Link %s Unexpected error' % link)
        else:
            print('Wrong type danger zone')
            errors.append('Link %s supplied link type %s is not valid - Please use physical or logical' % (link, link_type))
    if len(errors) > 0:
        raise Exception(errors)

def check_link_status(self,link,status='up',description=None):
    ''' Checks specified interface link is in the state desired.

    Check Link Status | link=ae1.0
    '''
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name

    output = self._vchannel.cmd("show interface %s" % link)
    if _check_if_output_empty(output) != True:
        errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)
    _check_link_status_state(output,status)
    if description != None:
      _check_link_status_description(output,description)
    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)

def get_chassis_full(self):
    ''' Gets the full output of the command 'show chassis hardware'

    Get Chassis Full
    '''
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name
    
    output = self._vchannel.cmd("show chassis hardware")
    if _check_if_output_empty(output) != True:
        errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)
    
    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)

    line = output.split('\n')[0]    # first line
    result=line
    tmp = line.split()
    if len(tmp) > 2:
        result = tmp[1]
    else:
        result = ''
    BuiltIn().log("\n   Got the chassis information: %s" % (result),console=True)
    return result

def get_route_summary(self,table='inet.0'):
    """ Returns number of active route in the ``table``

    ``table`` could be ``inet.0`` or ``inet.6``

    Get Route Summary | table=inet.0
    """
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name

    result = ""
    output = self._vchannel.cmd("show route summary table %s" % table)
    if _check_if_output_empty(output) != True:
        errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)
    
    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)
    
    for line in output.split("\n"):
        match = re.match(r"%s: .* \((.+) active," % table, line)
        if match:
            result = int(match.group(1))
    BuiltIn().log("Got %d routes from `%s`" % (result,table)) 
    return result 

def get_route_summary_txt(self,table='inet.0',file_out='default'):
    ''' 
    Returns number of active route in the ``table`` and saves as a file

    ``table`` could be ``inet.0`` or ``inet.6``

    Get Route Summary | table=inet.0 | fileout=route.txt
    '''
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name

    result = ''
    output = self._vchannel.cmd("show route summary table %s" % table)
    if _check_if_output_empty(output) != True:
        errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)

    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)

    for line in output.split("\n"):
        match = re.match(r"%s: .* \((.+) active," % table, line)
        if match:
            result = int(match.group(1))
            BuiltIn().log("Got %d routes from `%s`" % (result,table),console=True) 
        if 'default' not in file_out:
            f = open('%s.txt' % file_out,'w+')
            f.write(str(result))
            f.close
    return result 

def verify_commit_mode(self):
    ''' Verifies the commit mode of the router is set to synchronize

    Verify Commit Mode
    '''
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name

    output = self._vchannel.cmd('show configuration system commit synchronize')

    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)

    if _check_if_output_empty(output) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    if 'synchronize' in output:
        BuiltIn().log('\n   Commit Syncronized Is Enabled',console=True)
    else:
        raise Exception('Commit Syncronized Is Not Enabled')

def engine_switch_test(self,checkstatus=''):
    """ Verifies the number of route engines present within the device.

    If more than 1 is present attempts to connect to the back up route-engine to verify functionality
    Test will fail if >1 RE present and backup RE is not reachable. 
    
    check_status=True causes the test to ensure RE0 is master.

    Verify Route Engine
    Verify Route Engine | check_status=True
    """ 
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name

    output = self._vchannel.cmd("request routing-engine login backup")
    if _check_if_output_empty(output) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    output = self._vchannel.cmd("request routing-engine login master")
    if _check_if_output_empty(output) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    output = self._vchannel.cmd("request routing-engine login backup")
    if _check_if_output_empty(output) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)

    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)

def verify_route_engine(self,checkstatus=''):
    """ Verifies the number of route engines present within the device.

    If more than 1 is present attempts to connect to the back up route-engine to verify functionality
    Test will fail if >1 RE present and backup RE is not reachable. 
    
    check_status=True causes the test to ensure RE0 is master.

    Verify Route Engine
    Verify Route Engine | check_status=True
    """ 

    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name

    output = self._vchannel.cmd("show chassis routing-engine | grep slot | count")
    if _check_if_output_empty(output) != True:
        errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)
    slots = re.findall(r'\b(\d)',output)
    slots = int(slots[0])

    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)

    if checkstatus == 'True' or checkstatus == 'true':
        output = self._vchannel.cmd("show chassis routing-engine 0 | no-more")
        if _check_if_output_empty(output) != True:
            raise Exception('%s - Output returned empty - Please check configuration on device' % device)
        if _verify_route_engine_status(output) == False:
            raise Exception('RE0 is not the master')
        output = self._vchannel.cmd("request routing-engine login master")
        if _check_if_output_empty(output) != True:
            raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    if slots > 1:
        output = self._vchannel.cmd("request routing-engine login backup | no-more")
        if _check_if_output_empty(output) != True:
            raise Exception('%s - Output returned empty - Please check configuration on device' % device)
        if 'No route' in output:
            raise Exception('Unable to connect to backup RE')
        output = self._vchannel.cmd("request routing-engine login master")
        if _check_if_output_empty(output) != True:
            raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    output = self._vchannel.cmd("set cli idle-timeout 0")
    output = self._vchannel.cmd("set cli screen-length 0")



def verify_version(self,expectedversion='default'):
    """ Verifies that the router is the expected version

    If no argument is specified simply logs the current version to the console

    If argument is specified ensures that the version on the device matches the expected version
    
    """ 
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name

    output = self._vchannel.cmd("show version | no-more")
    if _check_if_output_empty(output) != True:
        errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)

    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)

    match = re.findall(r'Junos: (.*)',output)
    if match == None or match == False:
        raise Exception('No Valid Match Found')
    if expectedversion == 'default':
        BuiltIn().log('Junos Version Found. Version: %s' %match[0])
        return match[0]
    elif match[0] == expectedversion:
        BuiltIn().log('Junos Version Matches Expected: %s' %match[0])
    else:
        raise Exception('Version Suppled Does Not Match')

def check_isis_adjacency(self,link='default',level='0',state='default'):
    '''Checks the ISIS Adjacency Status of the specified interfaces.
    If no arguments are specified it will simply log the output to the console

    Currently assumes only deisred state is up

    Check ISIS Adjacency
    Check ISIS Adjacency | ae1
    '''
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name
    
    #Default command for tests without any options, Prevents unessecary failures whilst highlighting the error.
    if link == 'default':
       BuiltIn().log('No Link Specified, Executing Show All Command')
       output = self._vchannel.cmd('show isis adjacency')
       BuiltIn().log(output,console=True)
       #Validate the return output
       if _check_if_output_empty(output) != True:
            errors.append("You're a redneck heartbreak, Who's really bad at lying")
       else:
           return
    #If a link argument has been specified.           
    elif link != 'default':
       output = self._vchannel.cmd('show isis adjacency | grep %s' % link)
       #Validate the return output
       if _check_if_output_empty(output) != True:
            errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)
       else:
            #BuiltIn().log('stuff', console=True)
            link_state = re.search(r'\s(Up)\s', output)
            if re.search(r'\s(Up)\s', output) != None:
                return True
                #BuiltIn().log(' truestuff', console=True)
            else:
                if '!' in output:                    
                    errors.append('%s - %s Adjacency Missing IP Address - Check ISIS configuration on device' % (self._vchannel.current_name,link))
                if 'Down' in output:
                    errors.append('%s - %s Adjacency Down - Check Device' % (self._vchannel.current_name,link))
                if 'One-way' in output:
                    errors.append('%s - %s Adjacency In One-Way State - Check ISIS configuration on Device' % (self._vchannel.current_name,link))
                if 'Initializing' in output:
                    errors.append('%s - %s Adjacency Initializing' % (self._vchannel.current_name,link))

    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)

def check_isis_interface(self,link='default', level='0',state='enabled'):
    '''Checks the ISIS Interface Status for the specified level & DR type.

    Checks if the specified level is enabled or disabled based on arguments.

    Valid states: point to point, passive, disabled

    Check ISIS Interface | link=ae8.0 | level=2 | state=point to point
    '''
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name

    if link == 'default':
        BuiltIn().log('No Link Specified, Executing Show All Command')
        output = self._vchannel.cmd('show isis interface')
    elif link != 'default':
        output = self._vchannel.cmd('show isis interface %s' % link)
        if _check_if_output_empty(output) != True:
            raise Exception('%s - Returned Output Was Empty - Please check supplied variables' % device)
        #Make sure link level has been specified.
        if level == '0':
            raise Exception('%s -  No Level Specified. Please Check Test Script')
        #If link level has been specified. 
        if level != '0':
            configured_level = re.findall(r'\s\s(\d)\s\s', output)
            if level == '1':
                configured_l1_dr = re.findall(r'\d\s(Point to Point)|\d\s(Passive)|\d\s(Disabled)', output)
                if str(state).lower() in str(configured_l1_dr).lower():
                    return True
                else:
                    errors.append('%s - Configured state %s does not match expected %s' % (device,configured_l1_dr,state))
            elif level == '2':
                configured_l2_dr = re.findall(r'\s\s(Point to Point)|\s\s(Passive)|\s\s(Disabled)', output)
                if str(state).lower() in str(configured_l2_dr).lower():
                    return True
                else:
                    errors.append('%s - Configured state %s does not match expected %s' % (device,configured_l2_dr,state))
            else: 
                errors.append('%s - Configured Level %s is not a valid level, Please use 1 or 2' % (device,configured_level))
    
    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)    

def check_isis_level_configuration(self,level='0',state='enabled'):
    '''Checks the ISIS Level Configuration Level for the specified level.

    Checks if the specified level is enabled or disabled based on arguments.

    Check ISIS Level Configuration | level=1 | state=disabled
    Check ISIS Level Configuration | level=2 | state=enabled
    '''
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name
    
    #Default command for tests without any options, Prevents unessecary failures whilst highlighting the error.
    if level == '0':
        raise Exception('%s -  No Level Specified. Please Check Test Script' % device)
    #If a link argument has been specified.           
    if level != '0':
        output = self._vchannel.cmd('show configuration protocols isis level %s' % level)
        if 'Invalid numeric value' in output or 'is not within range' in output:
            raise Exception ('%s - %s is not a valid argument, Please use a value between 1-2' % (device, level))
        if state == 'enabled':
            if 'disable' not in output:
                return True
            else:
                errors.append('%s - State Does not match Desired, Got - Disabled, Expected: %s' %(device, state))
        elif state == 'disabled':
            if 'disable' in output:
                return True
            else:
                errors.append('%s - State Does not match Desired, Got - Enabled, Expected: Disabled' % device)
        else:        
            errors.append('%s - Invalid State Supplied - %s' %(device, state))
                        
    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors) 

def verify_isis_protocol(self,protocol='ipv4',state='enabled'):
    '''Checks the specified procotol matches the desired state for ISIS.

    Checks if the specified level is enabled or disabled based on arguments.

    If no argument supplied IPv4 is automatically checked.

    Verify ISIS Protocol | protocol=Ipv4 | state=enabled
    '''
    #Create errors for appending into
    errors = []
    device = self._vchannel.current_name
    
    if protocol.lower() == 'ipv4':
        output = self._vchannel.cmd('show isis overview | grep %s' % protocol)
        if _check_if_output_empty(output) != True:
            raise Exception('%s - Returned Output Was Empty - Please check supplied variables' % device)
        current_status = re.findall(r'IPv4 is (\w+)', output)   
        BuiltIn().log(output, console=True)
        BuiltIn().log(current_status[0], console=True)
        if state.lower() == current_status[0].lower():
            return True
        else:
            errors.append('%s - State Does not match Desired, Got - %s, Expected: %s' %(device,current_status,state))
    elif protocol.lower() == 'ipv6':
        output = self._vchannel.cmd('show isis overview | grep %s' % protocol)
        if _check_if_output_empty(output) != True:
            raise Exception('%s - Returned Output Was Empty - Please check supplied variables' % device)
        current_status = re.findall(r'IPv6 is (\w+)', output)   
        BuiltIn().log(output, console=True)
        BuiltIn().log(current_status[0], console=True)
        if state.lower() == current_status[0].lower():
            return True
        else:
            errors.append('%s - State Does not match Desired, Got - %s, Expected: %s' %(device,current_status,state))
    else:
        raise Exception('Invalid protocol specified - %s . Please use ipv4 or ipv6')    
                       
    #Check if any errors where raised as part of the process and raise them.
    if len(errors) > 0:
        raise Exception(errors)

def verify_isis_wide_metrics(self,level='0',state='enabled'):
    '''Checks for wide metrics being enabled for specified level for ISIS.

    Checks if the specified level is enabled or disabled based on arguments.

    If no state argument supplied enabled is automatically checked.

    Verify ISIS Wide Metrics | level=2 | state=enabled
    '''
    errors = []
    device = self._vchannel.current_name

    if level == '0':
        errors.append('%s - No level specified - Please Check Test Script' % device)
    if level not in ['1','2']:
        errors.append('%s - Level %s specified is not a valid ISIS level - Please Check Test Script' % (device,level))
    if state.casefold() not in ['enabled, disabled']:
        errors.append('%s -  %s not a valid state. Please use enabled or disabled.' % (device,state))
    
    #Raise errors if initial input sanity tests fail
    if len(errors) > 0:
        raise Exception(errors)
    
    #If a link argument has been specified.           
    if level != '0':
        if level == '1':
            output = self._vchannel.cmd('show isis overview | find "level 1" | grep wide')
            if _check_if_output_empty(output) != True:
                raise Exception('%s - Returned Output Was Empty - Please check supplied variables' % device)
            current_state = re.findall(r'are (\w+)',output)
            if current_state[0].casefold() == state.casefold():
                return True
            else:
                raise Exception('%s - State Does not match Desired, Got - %s, Expected: %s' %(device,current_state,state))          
        elif level == '2':
            output = self._vchannel.cmd('show isis overview | find "level 2" | grep wide')
            if _check_if_output_empty(output) != True:
                raise Exception('%s - Returned Output Was Empty - Please check supplied variables' % device)  
            current_state = re.findall(r'are (\w+)',output)
            if current_state[0].casefold() == state.casefold():
                return True
            else:
                raise Exception('%s - State Does not match Desired, Got - %s, Expected: %s' %(device,current_state,state))                    


def verify_isis_link_costs(self,link='default', level='0',metric='enabled'):
    '''Checks the specified selected link has the desired metrics state for ISIS.

    Validates input for a link name, valid level and numerical metric.

    All arguments required 

    Verify ISIS Link Costs | link=ae8.0  | level=2 | metric=10
    '''
    errors = []
    device = self._vchannel.current_name

    if link == 'default':
        errors.append('%s No Link Specified, Please Check Test Script' % device)
    elif link != 'default':
        output = self._vchannel.cmd('show isis interface %s' % link)
        if _check_if_output_empty(output) != True:
            errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)  
    if level == '0':
        errors.append('%s - No level specified - Please Check Test Script' % device)
    elif level not in ['1','2']:
        errors.append('%s - Level %s specified is not a valid ISIS level - Please Check Test Script' % (device,level))

    if metric== 'enabled':
        errors.append('%s - No metric specified - Please Check Test Script' % device) 
    elif metric.isdigit() == False:
        errors.append('%s - Metric specified %s is not a number - Please Check Test Script' % (device,metric))           
    
    #Raise errors if initial input sanity tests fail
    if len(errors) > 0:
        raise Exception(errors)

    if level == '1':
        current_metric = re.findall(r'(\d+)\/\d+',output)
        if int(current_metric[0]) == int(metric):
            return True
        elif int(current_metric[0]) != int(metric):
            errors.append('%s - Metric Does not match Desired, Got - %s, Expected: %s' %(device,current_metric[0],metric))                    
    if level == '2':
        current_metric = re.findall(r'\d+\/(\d+)',output)
        if int(current_metric[0]) == int(metric):
            return True
        elif int(current_metric[0]) != int(metric):
            errors.append('%s - Metric Does not match Desired, Got - %s, Expected: %s' %(device,current_metric[0],metric))                    

    if len(errors) > 0:
        raise Exception(errors)

def verify_isis_adjacency_authentication(self,link='default',level='0',authentication='md5',key_chain='iih'):
    '''Checks the specified link matches md5 enabled authentication has the for ISIS.

    Verify ISIS Adjacency Authentication | link=ae8 | level=2
    '''
    errors = []
    device = self._vchannel.current_name

    if link == 'default':
        BuiltIn().log('%s No Link Specified, Please Check Test Script' % device)
        errors.append('%s No Link Specified, Please Check Test Script' % device)
    elif link != 'default':
        output = self._vchannel.cmd('show isis authentication | grep %s' % link)
        if _check_if_output_empty(output) != True:
            errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)  
    if level == '0':
        errors.append('%s - No level specified - Please Check Test Script' % device)
    elif level not in ['1','2']:
        errors.append('%s - Level %s specified is not a valid ISIS level - Please Check Test Script' % (device,level))

    #Raise errors if initial input sanity tests fail
    if len(errors) > 0:
        raise Exception(errors) 

    #Capture link level
    isis_link_level = re.findall(r'\s\s(1|2)\s\s',output)
    #Capture Key Chains, All not used currently. - Kept for future expansion potentials.
    iih_auth_chain = re.findall(r'(\w+)\s\w+\s\w+',output)
    csn_auth_chain = re.findall(r'\w+\s(\w+)\s\w+',output)
    psn_auth_chain = re.findall(r'\w+\s\w+\s(\w+)',output)


    #Must be a neater way to do this
    if key_chain == 'iih':
        output = self._vchannel.cmd('show security keychain detail | find %s' % iih_auth_chain[0])
        if _check_if_output_empty(output) != True:
            errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)  
        if isis_link_level[0] == level:
            keychain_algorithm = psn_auth_chain = re.findall(r'Algorithm\s(.*),\sState',output)
            BuiltIn().log(keychain_algorithm[0],console=True)
            if authentication.casefold() in keychain_algorithm[0].casefold():
                return True
            else:
                errors.append('%s - Configured authentication %s does not match expected authentication %s - Please Check Configuration' % (device,authentication,keychain_algorithm[0]))
        elif isis_link_level[0] != level:
            errors.append('%s - Configured level %s does not match expected level %s - Please Check Configuration' % (device,isis_link_level[0],level))
    if key_chain == 'csn':
        output = self._vchannel.cmd('show security keychain detail | find %s' % csn_auth_chain[0])
        if _check_if_output_empty(output) != True:
            errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)  
        if isis_link_level[0] == level:
            keychain_algorithm = psn_auth_chain = re.findall(r'Algorithm\s(.*),\sState',output)
            BuiltIn().log(keychain_algorithm[0],console=True)
            if authentication.casefold() in keychain_algorithm[0].casefold():
                return True
            else:
                errors.append('%s - Configured authentication %s does not match expected authentication %s - Please Check Configuration' % (device,authentication,keychain_algorithm[0]))
        elif isis_link_level[0] != level:
            errors.append('%s - Configured level %s does not match expected level %s - Please Check Configuration' % (device,isis_link_level[0],level))
    if key_chain == 'psn':
        output = self._vchannel.cmd('show security keychain detail | find %s' % psn_auth_chain[0])
        if _check_if_output_empty(output) != True:
            errors.append('%s - Returned Output Was Empty - Please check supplied variables.' % device)  
        if isis_link_level[0] == level:
            keychain_algorithm = psn_auth_chain = re.findall(r'Algorithm\s(.*),\sState',output)
            BuiltIn().log(keychain_algorithm[0],console=True)
            if authentication.casefold() in keychain_algorithm[0].casefold():
                return True
            else:
                errors.append('%s - Configured authentication %s does not match expected authentication %s - Please Check Configuration' % (device,authentication,keychain_algorithm[0]))
        elif isis_link_level[0] != level:
            errors.append('%s - Configured level %s does not match expected level %s - Please Check Configuration' % (device,isis_link_level[0],level))

    #Raise errors produced via testing
    if len(errors) > 0:
        raise Exception(errors)

def verify_ldp_neighbour(self,link='default',neighbour='default',notenabled='default'):
    '''Verifies that the expected neighbour is seen at the other side of an LDP link,

    Supports alternative of validating that a neighbour does not exist with the use of 'notenabled=true'

    Verify LDP Neighbour | link=ae8.0 | neighbour=127.0.0.1
    Verify LDP Neighbour | link=ae8.0 | neighbour=127.0.0.1| notenabled=true
    '''
    errors = []
    device = self._vchannel.current_name
    if link == 'default':
        errors.append('%s No Link Specified, Please Check Test Script' % device)
    if neighbour == 'default':
        errors.append('%s No Neighbour Specified, Please Check Test Script' % device)
    if neighbour != 'default':
        try:
            ip_address(neighbour) 
        except:
            errors.append('%s - %s not a valid IPv4 Address, Please use Enabled or Disabled' %(device,neighbour))   

    #Raise errors if initial input sanity tests fail
    if len(errors) > 0:
        raise Exception(errors) 

    if notenabled != 'default':
        output = self._vchannel.cmd('show ldp neighbor | grep %s' % link)
        if _check_if_output_empty(output) != True:
            return True   
        else:
            errors.append('%s - Neighbour %s in connected to  %s, Please check configuration' %(device, neighbour, link))     
    else:
        output = self._vchannel.cmd('show ldp neighbor | grep %s' % link)
        if _check_if_output_empty(output) != True:
            raise Exception('%s - Returned Output Was Empty - Please check supplied variables.' % device)  
        if neighbour not in output:
            errors.append('%s - Neighbour %s not connected to link %s, Please check configuration' %(device, neighbour, link))

    #Raise errors if tests fail
    if len(errors) > 0:
        raise Exception(errors)

def verify_ldp_session(self,neighbour='default',state='operational',notenabled='default'):
    '''Verifies that the expected neighbour session is seen at the other side of an LDP link,

    Assumes the default state of operational if no other specified.

    Supports alternative of validating that a neighbour does not exist with the use of 'notenabled=true'

    Verify LDP Session | neighbour=127.0.0.1 | state=enabled
    Verify LDP Session | neighbour=127.0.0.1 | state=enabled | notenabled=true
    '''
    errors = []
    device = self._vchannel.current_name

    try:
        ip_address(neighbour)
    except:
        errors.append('Neighbour address provided not a valid IPV4 address')
    # if state == 'default':
    #     errors.append('%s No State Specified, Please Check Test Script' % device)
    #Raise errors if initial input sanity tests fail
    if len(errors) > 0:
        raise Exception(errors) 

    current_state = re.findall(r'\s\s(\w+)\s\s\w+\s',output)
    current_status = re.findall(r'\s\s\w+\s\s(\w+)\s',output)

    if notenabled != 'default':
        output = self._vchannel.cmd(' show ldp session %s' % neighbour)
        if _check_if_output_empty(output) != True:
            return True   
        else:
            errors.append('%s - Neighbour %s in state %s, Please check configuration' %(device, neighbour, current_state))     
    else:
        output = self._vchannel.cmd(' show ldp session %s' % neighbour)
        if _check_if_output_empty(output) != True:
            raise Exception('%s - Returned Output Was Empty - Please check supplied variables.' % device)  
        if current_status[0].casefold() in ['closed','opening']:
            errors.append('%s - Neighbour %s in connection state %s, Please check configuration' %(device, neighbour, current_state))
        if current_state[0].casefold() in ['Nonexistent', 'Connecting', 'Initialized', 'OpenRec', 'OpenSent', 'Closing']:
            errors.append('%s - Neighbour %s in state %s, Please check configuration' %(device, neighbour, current_state))
        if current_state[0].casefold() == state:
            return True
        if current_state[0].casefold() != state:
            errors.append('%s - Neighbour %s is not in the Expected State: %s - Expected: %s, Please check configuration' %(device, neighbour, current_state[0],state))

    #Raise errors if tests fail
    if len(errors) > 0:
        raise Exception(errors)

def verify_ldp_hellos(self,neighbour='default',state='enabled'):
    '''Verifies that the expected neighbour session has link protection enabled,

    Will accept arguments of state=enabled/disabled dependent on desired state

    Verify LDP Hellos | neighbour=127.0.0.1 | state=enabled
    '''
    errors = []
    device = self._vchannel.current_name

    try:
        ip_address(neighbour) 
    except:
        errors.append('%s - %s not a valid IPv4 Address, Please use Enabled or Disabled' %(device,neighbour))
    if state.casefold() not in ['enabled', 'disabled']:
        errors.append('%s - State %s not a valid protection state, Please use Enabled or Disabled' %(device,state))
    if len(errors) > 0:
        raise Exception(errors) 

    output = self._vchannel.cmd(' show ldp session extensive %s | grep protection' % neighbour)
    if _check_if_output_empty(output) != True:
        raise Exception('%s - Returned Output Was Empty - Please check supplied variables.' % device)  

    #neighbour_type = re.findall(r'Neighbor types: (.*)',output)        Might factor into later tests
    current_state = re.findall(r'Protection: (\w+)',output)

    if current_state[0].casefold() == state:
        return True
    if current_state[0].casefold() != state:
        errors.append('%s - Neighbour %s protection is not in the Expected State - Current State: %s - Expected: %s - Please check configuration' %(device, neighbour, current_state[0],state))

    #Raise errors if tests fail
    if len(errors) > 0:
        raise Exception(errors)

def verify_ldp_igp_tracking(self,instance='default'):
    '''Verifies that igp metric tracking is enabled for LDP on the router/Instance

    Assumes no routing instance if no instance specified.

    Verify LDP IGP Tracking
    Verify LDP IGP Tracking | instance=customerinstance
    '''
    errors = []
    device = self._vchannel.current_name

    if instance != 'default':
        output = self._vchannel.cmd('show configuration %s' % instance)
        if _check_if_output_empty(output) != True:
            raise Exception('%s - No such routing instance %s - Please check supplied variables.' % (device,instance))  

    if instance == 'default':
        output = self._vchannel.cmd('show configuration protocols ldp | find track-igp-metric')
        if _check_if_output_empty(output) != True:
            raise Exception('%s - Returned Output Was Empty - Please check supplied variables.' % device)  
        if 'track-igp-metric;' in output:
            return True
        else:
            errors.append('%s - IGP Metric Tracking Not Enabled' % device)
    else:
        output = self._vchannel.cmd('show configuration routing-instances %s protocols ldp | find track-igp-metric' % instance)
        if _check_if_output_empty(output) != True:
            raise Exception('%s - Returned Output Was Empty - Please check supplied variables.' % device)  
        if 'track-igp-metric;' in output:
            return True
        else:
            errors.append('%s - IGP Metric Tracking Not Enabled' % device)

    #Raise errors if tests fail
    if len(errors) > 0:
        raise Exception(errors)

def verify_bgp_peer_address_family(self, instance='default',family='inet-vpn',peer_group='default'):
    '''Verifies that specified BGP families are configured on peer group

    Validates family type - Acceptable 'inet','inet6','inet-vpn','inet6-vpn','iso-vpn'

    Does not require instance, Defaults to default instance if no instance specified.

    Verify BGP Peer Address Family | peer_group=V4_IBGP  | family=inet-vpn
    Verify BGP Peer Address Family | instance=customerinstance | peer_group=V4_IBGP  | family=inet-vpn

    '''
    errors = []
    device = self._vchannel.current_name

    if family not in ['inet','inet6','inet-vpn','inet6-vpn','iso-vpn']:
        errors.append('%s - %s Family not a valid BGP family - Please check supplied variables.' % (device,family))
    if instance != 'default':
        output = self._vchannel.cmd('show configuration routing-instances %s' % instance)
        if _check_if_output_empty(output) != True:
            errors.append('%s - %s Family not a valid BGP family - Please check supplied variables.' % (device,family))       
    #Raise errors if initial input sanity tests fail
    if len(errors) > 0:
        raise Exception(errors) 

    family_string = 'family '+family
    BuiltIn().log(family_string,console=True)

    if instance == 'default':
        output = self._vchannel.cmd('show configuration protocols bgp group %s' % peer_group)
        if _check_if_output_empty(output) != True:
           raise Exception('%s - Output Returned Empty - Please check peer group.' % device)
        if family_string in output:
            return True
        else:
            errors.append('%s - %s Family not present in group %s - Please check supplied variables.' % (device,family, peer_group))       
    else:
        output = self._vchannel.cmd('show configuration routing-instances %s protocols bgp group %s' % (instance,peer_group))
        if _check_if_output_empty(output) != True:
           raise Exception('%s - Output Returned Empty - Please check peer group.' % device)
        if family_string in output:
            return True
        else:
            errors.append('%s - %s Family not present in %s group %s - Please check supplied variables.' % (device,family,instance,peer_group))       
    
    #Raise errors if tests fail
    if len(errors) > 0:
        raise Exception(errors)
    
    
def verify_bgp_local_address(self, address='default',link='default',group='default',instance='default'):
    '''Verifies that specified BGP peer group has desired local address.

    Does not require instance, Defaults to default instance if no instance specified.

    Verify BGP Local Address | group=V4_IBGP | address=127.0.0.1 | 
    Verify BGP Local Address | instance=customer instance | group=V4_IBGP  | address=127.0.0.1 | 

    '''
    errors = []
    device = self._vchannel.current_name

    if group == 'default':
        errors.append('%s No Group Specified, Please Check Test Script' % device)
    if address != 'default':
        try:
            ip_address(address) 
        except:
            errors.append('%s - %s not a valid IPv4 Address, Please use Enabled or Disabled' %(device,address))
    if instance != 'default':
        output = self._vchannel.cmd('show configuration routing-instances %s' % instance)
        if _check_if_output_empty(output) != True:
            errors.append('%s - %s is not a valid instance - Please check supplied variables.' % (device,instance))
    
    #Raise errors if initial input sanity tests fail
    if len(errors) > 0:
        raise Exception(errors) 


    if instance == 'default':
        output = self._vchannel.cmd('show configuration protocols bgp group %s' % group)
        if _check_if_output_empty(output) != True:
            errors.append('%s - %s is not a valid group - Please check supplied variables.' % (device,instance))
        local_address = re.findall(r'local-address (.*);',output)
        if len(local_address) == 0:
            raise Exception('%s - Could not find local address for link %s'% (device,link))
        BuiltIn().log(local_address,console=True)
        BuiltIn().log(address,console=True)

        if local_address[0] == address:
            BuiltIn().log('southen girl rock my world',console=True)
            return True
        elif local_address[0] != address:
            BuiltIn().log('stuck like glue',console=True)
            errors.append('%s - %s does not match current group %s address %s - Please check supplied variables.' % (device,address,group,local_address))
    elif instance != 'default':
        output = self._vchannel.cmd('show configuration routing-instances %s protocols bgp group %s' % (instance,group))
        if _check_if_output_empty(output) != True:
            errors.append('%s - %s is not a valid instance - Please check supplied variables.' % (device,instance))
        local_address = re.findall(r'local-address (.*);',output)
        if len(local_address) == 0:
            raise Exception('%s - Could not find local address for link %s'% (device,link))
        BuiltIn().log(local_address,console=True)
        BuiltIn().log(address,console=True)
        if local_address[0] == address:
            BuiltIn().log('southen girl rock my world',console=True)
            return True
        elif local_address[0] != address:
            BuiltIn().log('stuck like glue',console=True)
            errors.append('%s - %s does not match current group %s address %s - Please check supplied variables.' % (device,address,group,local_address))
    
    #Raise errors if tests fail
    if len(errors) > 0:
        raise Exception(errors)

def verify_ntp_configuration(self):
    '''Verifies that existence but not VALIDITY of NTP configuration on the router.

    Verify NTP Configuration 

    '''
    output = self._vchannel.cmd('show configuration system ntp')
    if _check_if_output_empty(output) != True:
        raise Exception('No NTP configuration on device')

def verify_ntp_synchronisation(self):
    '''Verifies that NTP is currently in sync with a suitable server

    Verify NTP Syncronisation
    '''
    device = self._vchannel.current_name

    output = self._vchannel.cmd('show ntp status')
    if _check_if_output_empty(output) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)

    if 'timed out' in output:
        raise Exception('%s - Request Timed out, Please check device configuration' % device)
    else:
        output_lines = output.splitlines()
        for line in output_lines[1:]:
            reach = re.findall(r'\d+\s\d+\s\s(\d+)\s+\d+.',line)
            if reach[0] == '0':
                raise Exception('%s - Device has 0 sucessful syncs with NTP server' % device)
            else:
                n = 0
                octal_value = ['1','3','7','17','37','77','177','377']
                for value in octal_value:
                    n = n+1
                    if reach[0] == str(n):
                        BuiltIn().log('%s - Device has had %s successful syncs' % (device,n),console=True)

def verify_secure_access_protocols(self):
    '''Verifies that only secure access methods are configured to the device. Raises error on the presence of insecure methods

    Errors searched for - 'finger','telnet','tftp-server'

    Verify Secure Access Protocols
    '''
    errors = []
    device = self._vchannel.current_name

    output = self._vchannel.cmd('show configuration system services')
    if _check_if_output_empty(output) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    trigger_words = ['finger','telnet','tftp-server']
    for trigger in trigger_words:
        if trigger in output:
            errors.append('%s - Device configured for %s - Please review device configuration' % (device, trigger))

    #Raise errors if tests fail
    if len(errors) > 0:
        raise Exception(errors)

def verify_syslog(self):
    '''Verifies that existence but not VALIDITY of NTP configuration on the router.

    Errors searched for - 'core dump','error','failure','critical','exception'

    Verify Syslog
    '''
    errors = []
    device = self._vchannel.current_name

    output = self._vchannel.cmd('show log messages | last | no-more')
    if _check_if_output_empty(output) != True:
        raise Exception('%s - No log messages present - Please check configuration on device' % device)

    BuiltIn().log('\n',console=True)
    trigger_words =  ['core dump','error','failure','critical','exception']        
    for trigger in trigger_words:
        output = self._vchannel.cmd('show log messages | no-more | except UI_CMDLINE_READ_LINE | grep "%s"' % trigger)
        if trigger in output:
            errors.append('%s - Device sys log contains %s errors - Please review device sys log' % (device, trigger))

    #Raise errors if tests fail
    if len(errors) > 0:
        raise Exception(errors)

def verify_boot_process(self):
    '''Verifies that existence the system has booted without any errors in the boot log

    Errors searched for - 'core dump','error','failure','critical','exception'

    Verify Boot Process

    '''
    errors = []
    device = self._vchannel.current_name

    output = self._vchannel.cmd('show system boot-messages')
    if _check_if_output_empty(output) != True:
        raise Exception('%s - Output returned empty - Please check configuration on device' % device)
    trigger_words = ['core dump','error','failure','critical','exception']
    for trigger in trigger_words:
        if trigger in output:
            errors.append('%s - Device sys log contains %s errors - Please review device sys log' % (device, trigger))

    #Raise errors if tests fail
    if len(errors) > 0:
        raise Exception(errors)

def verify_ftp(self,source='default',destination='default'):
    '''Verifies FTP functionality by attempting to copy to or from a specified location

    Verify FTP | source=ftp://ftp:nobody@127.0.0.1/test.txt | destination=test.txt
    '''
    errors = []
    device = self._vchannel.current_name

    if source == 'default':
        errors.append('%s No source file specified, Please Check Test Script' % device)
    if destination == 'default':
        errors.append('%s No destination file specified, Please Check Test Script' % device)
    
    #Raise errors if initial input sanity tests fail
    if len(errors) > 0:
        raise Exception(errors) 
    output = self._vchannel.cmd('file copy %s %s' % (source,destination))
    if 'error:' in output:
        errors.append('%s Error in command , Please Check Test Script' % device)
        errors.append(output)
    elif 'transferring' in output:
        return True

    #Raise errors if initial input sanity tests fail
    if len(errors) > 0:
        raise Exception(errors) 
